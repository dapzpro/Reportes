import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("Filtrar Reportes de Carros")

# Subir archivo
uploaded_file = st.file_uploader("Sube el archivo de reporte (.xlsx)", type=["xlsx"])

if uploaded_file:
    # Leer archivo con pandas asegurando que no se pierdan filas
    informe = pd.read_excel(uploaded_file, engine="openpyxl")
    st.write(f"El archivo original tiene {informe.shape[0]} filas.")


    # Filtrar alertas que no sean NaN
    alertas = informe[informe["Alerta"].notna()].copy()

    # Convertir columna "Hora" a formato datetime (solo hora)
    alertas['Hora'] = pd.to_datetime(alertas['Hora'], format='%H:%M:%S').dt.time

    # Para poder calcular diferencias, convertimos 'Hora' a timedelta
    alertas['Hora_Timedelta'] = pd.to_timedelta(alertas['Hora'].astype(str))

    # Ordenar por Hora en orden ascendente
    alertas = alertas.sort_values('Hora_Timedelta', ascending=True)

    # Calcular diferencia entre horas
    alertas['Diferencia'] = alertas['Hora_Timedelta'].diff()

    # Filtrar inconsistencias: diferencias menores a 1 segundo
    inconsistencias = alertas[alertas['Diferencia'].dt.total_seconds() < 1]

    # Incluir la fila anterior de las inconsistencias
    inconsistencias = pd.concat([alertas.iloc[1:][alertas['Diferencia'].dt.total_seconds() < 1], inconsistencias])


    # Eliminar las filas de inconsistencias del DataFrame 'alertas'
    alertas_limpias = alertas[~alertas.index.isin(inconsistencias.index)]

    # Eliminar las columnas 'Hora_Timedelta' y 'Diferencia' antes de la descarga
    alertas_limpias = alertas_limpias.drop(columns=['Hora_Timedelta', 'Diferencia'])

    # Agregar un selector para que el usuario elija qué columna filtrar
    opciones_columnas = alertas.columns.tolist()  # Listar todas las columnas disponibles
    columna_filtro = st.selectbox("¿Por qué columna te gustaría filtrar las alertas?", opciones_columnas)

    # Agregar un selector para filtrar por valor de la columna elegida
    valores_filtro = alertas[columna_filtro].unique().tolist()
    valor_filtro = st.selectbox(f"Selecciona el valor para filtrar en la columna '{columna_filtro}'", valores_filtro)

    # Filtrar las alertas según la opción seleccionada
    alertas_filtradas = alertas[alertas[columna_filtro] == valor_filtro]

    # Mostrar las alertas filtradas
    st.subheader(f"Alertas filtradas por la columna '{columna_filtro}' y valor '{valor_filtro}'")
    st.write(alertas_filtradas)

    # Mostrar las inconsistencias
    st.subheader("Inconsistencias encontradas (menos de 1 segundo)")
    st.write(inconsistencias)

    # Botón de "Más Opciones"
    with st.expander("Más Opciones"):
        # Contar cuántas coincidencias encontró
        num_coincidencias = alertas_filtradas.shape[0]
        st.write(f"Se encontraron {num_coincidencias} coincidencias para el filtro seleccionado.")

        # Mostrar estadísticas adicionales (por ejemplo, estadísticas de la columna 'Hora' o cualquier otra)
        # Aquí puedes agregar otras métricas útiles
        st.write("Estadísticas de las alertas filtradas:")
        st.write(alertas_filtradas.describe())

    # Botón desplegable para mostrar el conteo de inconsistencias
    with st.expander("Ver conteo de inconsistencias"):
        num_inconsistencias = inconsistencias.shape[0]
        num_filas_eliminadas = num_inconsistencias * 2
        st.write(f"Se encontraron {num_inconsistencias} inconsistencias, lo que resultó en la eliminación de {num_filas_eliminadas} filas.")

    # Crear archivo de Excel para alertas limpias
    excel_file = io.BytesIO()
    alertas_limpias.to_excel(excel_file, index=False, engine='openpyxl')
    excel_file.seek(0)  # Asegurarse de que el archivo esté listo para descarga


    st.write(f"El archivo sin inconsistencias tiene {alertas_limpias.shape[0]} filas.")

    # Botón para descargar el archivo de alertas limpias
    st.download_button(
        label="Descargar alertas sin inconsistencias 📄",
        data=excel_file,
        file_name="alertas_limpias.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Descargar archivo con inconsistencias marcadas
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Usar el DataFrame original para marcar las inconsistencias
    df_marcado = informe.drop(columns=["Hora_Timedelta", "Diferencia"], errors='ignore')
    
    for r in dataframe_to_rows(df_marcado, index=False, header=True):
        ws.append(r)

    # Identificar las inconsistencias en el DataFrame original (informe)
    inconsistentes_idx = inconsistencias.index.tolist()
    
    # Marcar las inconsistencias en el archivo Excel
    for i, idx in enumerate(df_marcado.index, start=2):  # Desde la fila 2 (por encabezado)
        if idx in inconsistentes_idx:
            for cell in ws[i]:
                cell.fill = red_fill

    archivo_marcado = io.BytesIO()
    wb.save(archivo_marcado)
    archivo_marcado.seek(0)

    st.download_button(
        label="Descargar reporte con inconsistencias marcadas 🔴",
        data=archivo_marcado,
        file_name="alertas_marcadas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
